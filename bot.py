import datetime
from openpyxl import Workbook, load_workbook
from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import Updater, CommandHandler, CallbackQueryHandler, MessageHandler, Filters, CallbackContext
import random

TOKEN = "8510356081:AAFrHnmw9ui7iQ-y4ADcEcXl1Tbd35rt8Eo"
ADMIN_CHAT_ID = 6125907347  

waiting_for_feedback = False


def save_feedback_to_excel(username, feedback):
    file_name = "feedback.xlsx"

    try:
        workbook = load_workbook(file_name)
        sheet = workbook.active
    except:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Username", "Feedback", "Date", "Time"])

    now = datetime.datetime.now()
    sheet.append([username, feedback, now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S")])
    workbook.save(file_name)


def start(update: Update, context: CallbackContext):
    keyboard = [
        [InlineKeyboardButton("🏖 Holidays", callback_data="holidays")],
        [InlineKeyboardButton("💡 Motivation", callback_data="motivation")],
        [InlineKeyboardButton("📅 Full Timetable", callback_data="timetable")],
        [InlineKeyboardButton("📝 Feedback", callback_data="feedback")]
    ]

    update.message.reply_text(
        "Hello! I am your KL Study Assistant Bot 🤖\n\nChoose what you need:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


def button_handler(update: Update, context: CallbackContext):
    global waiting_for_feedback
    query = update.callback_query
    query.answer()

    if query.data == "holidays":
        query.edit_message_text(
            "🏖 *KL University – Public Holidays 2025*\n\n"
            "1. *13-01-2025 to 15-01-2025* (Mon–Wed)\n"
            "   Pongal Holidays / Bhogi / Sankranthi / Kanuma\n\n"
            "2. *26-02-2025* (Wed)\n"
            "   Maha Shivaratri\n\n"
            "3. *31-03-2025* (Mon)\n"
            "   Ramzan (Id–Ul–Fitr)\n\n"
            "4. *08-08-2025* (Fri)\n"
            "   Varalakshmi Vratham\n\n"
            "5. *16-08-2025* (Sat)\n"
            "   Krishnashtami\n\n"
            "6. *27-08-2025* (Wed)\n"
            "   Vinayaka Chavithi\n\n"
            "7. *29-09-2025 to 04-10-2025* (Mon–Sat)\n"
            "   Dussehra Vacation\n\n"
            "8. *20-10-2025 & 21-10-2025* (Mon & Tue)\n"
            "   Deepavali\n\n"
            "9. *25-12-2025* (Thu)\n"
            "   Christmas",
            parse_mode="Markdown"
        )

    elif query.data == "motivation":
        tips = [
            "🔥 Success doesn't come from what you do occasionally, but from what you do consistently!",
            "📚 Study while others are sleeping. Dream while others are wishing.",
            "💪 Don’t stop when you're tired. Stop when you are done!",
            "🎯 Small daily progress adds up to big results.",
            "⏳ 45 minutes of focused study > 3 hours of distracted study!",
            "🚫 Stop doubting yourself — work hard and make it happen."
        ]
        query.edit_message_text("💡 *Daily Motivation*\n\n" + random.choice(tips), parse_mode="Markdown")

    elif query.data == "timetable":
        query.edit_message_text(
        "📅 *KL University Daily Timetable*\n\n"
        "⏰ *Regular Class Slots*\n\n"
        "1️⃣ 07:10 – 08:00 (50 min)\n"
        "2️⃣ 08:00 – 08:50 (50 min)\n\n"
        "🧋 *Break*: 08:50 – 09:20 (30 min)\n\n"
        "3️⃣ 09:20 – 10:10 (50 min)\n"
        "4️⃣ 10:10 – 11:00 (50 min)\n\n"
        "☕ *Break*: 11:00 – 11:10 (10 min)\n\n"
        "5️⃣ 11:10 – 12:00 (50 min)\n"
        "6️⃣ 12:00 – 12:50 (50 min)\n\n"
        "🍽 *Lunch Breaks*\n"
        "   Phase 1: 12:00 – 12:50\n"
        "   Phase 2: 12:50 – 13:50\n"
        "   Phase 3: 13:50 – 14:50\n\n"
        "7️⃣ 13:00 – 13:50 (50 min)\n"
        "8️⃣ 13:50 – 14:40 (50 min)\n\n"
        "🧋 *Break*: 14:40 – 14:50 (10 min)\n\n"
        "9️⃣ 14:50 – 15:40 (50 min)\n"
        "🔟 15:50 – 16:40 (50 min)\n"
        "1️⃣1️⃣ 16:40 – 17:30 (50 min)\n",
        parse_mode="Markdown"
    )


    elif query.data == "feedback":
        waiting_for_feedback = True
        query.edit_message_text(
            "📝 Please type your feedback below.\nYour opinion helps us improve this bot 😊"
        )


def text_handler(update: Update, context: CallbackContext):
    global waiting_for_feedback

    if waiting_for_feedback:
        feedback = update.message.text
        waiting_for_feedback = False
        username = update.message.from_user.username or "Anonymous"

        save_feedback_to_excel(username, feedback)

        try:
            context.bot.send_message(
                chat_id=ADMIN_CHAT_ID,
                text=f"📩 New Feedback:\n\n{feedback}\n\nFrom: @{username}",
                parse_mode="Markdown"
            )
        except:
            print("Failed to send to admin")

        update.message.reply_text(
            "🙏 Thank you for your feedback! We’ll review it and try to implement improvements soon 🚀",
            parse_mode="Markdown"
        )
    else:
        update.message.reply_text("Use /start to open menu 🙂")


def main():
    updater = Updater(TOKEN, use_context=True)
    dp = updater.dispatcher

    dp.add_handler(CommandHandler("start", start))
    dp.add_handler(CallbackQueryHandler(button_handler))
    dp.add_handler(MessageHandler(Filters.text, text_handler))

    updater.start_polling()
    updater.idle()


if __name__ == "__main__":
    main()
